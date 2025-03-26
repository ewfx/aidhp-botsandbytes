import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def create_excel_template():
    # Create a new Excel workbook
    wb = Workbook()
    
    # Users sheet
    ws_users = wb.active
    ws_users.title = "Users"
    ws_users['A1'] = "username"
    ws_users['B1'] = "interests"
    ws_users.append(["john_doe", "technology,programming,AI"])
    ws_users.append(["jane_smith", "health,fitness,cooking"])
    
    # Content sheet
    ws_content = wb.create_sheet("Content")
    ws_content['A1'] = "title"
    ws_content['B1'] = "category"
    ws_content['C1'] = "tags"
    ws_content.append(["Machine Learning Basics", "Technology", "AI,programming,data science"])
    ws_content.append(["Healthy Recipes", "Food", "cooking,health,nutrition"])
    
    # Interactions sheet
    ws_interactions = wb.create_sheet("Interactions")
    ws_interactions['A1'] = "username"
    ws_interactions['B1'] = "content_title"
    ws_interactions['C1'] = "interaction_type"
    ws_interactions['D1'] = "timestamp"
    ws_interactions.append(["john_doe", "Machine Learning Basics", "view", "2025-03-24 10:00:00"])
    ws_interactions.append(["jane_smith", "Healthy Recipes", "like", "2025-03-24 11:00:00"])
    
    # Style headers
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    header_font = Font(bold=True)
    
    for ws in [ws_users, ws_content, ws_interactions]:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
    
    # Save the template
    template_path = "templates/import_template.xlsx"
    wb.save(template_path)
    print(f"Template created at {template_path}")

if __name__ == "__main__":
    import os
    if not os.path.exists("templates"):
        os.makedirs("templates")
    create_excel_template()
